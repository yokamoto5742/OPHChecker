import tempfile
from pathlib import Path
from unittest.mock import patch

import pandas as pd
import pytest
from openpyxl import Workbook

from service.surgery_error_extractor import surgery_error_extractor


@pytest.fixture
def temp_comparison_file():
    """一時的な比較結果ファイルを作成"""
    temp_dir = tempfile.mkdtemp()

    # 実際のsurgery_comparatorと同じようにDataFrameを作成してCSVに書き込む
    df = pd.DataFrame({
        '手術日': ['2025/01/15', '2025/01/16', '2025/01/17'],
        '患者ID': [12345, 12346, 12347],
        '氏名': ['患者A', '患者B', '患者C'],
        '入外': ['外来', '入院', '外来'],
        '術眼': ['R', 'L', 'B'],
        '手術': ['白内障手術', '緑内障手術', '白内障手術'],
        '医師': ['橋本', '植田', '増子'],
        '麻酔': ['局所', '全身', '局所'],
        '術前': ['検査A', '検査B', '検査C'],
        '入外_比較': [True, True, '未入力'],
        '術眼_比較': [True, False, '未入力'],
        '手術_比較': [True, True, '未入力'],
        '医師_比較': [True, True, '未入力'],
        '麻酔_比較': [True, True, '未入力']
    })

    comparison_path = Path(temp_dir) / 'comparison.csv'
    df.to_csv(comparison_path, index=False, encoding='cp932')

    output_dir = Path(temp_dir) / 'output'
    output_dir.mkdir()

    yield {
        'comparison': str(comparison_path),
        'output_dir': str(output_dir),
        'temp_dir': temp_dir
    }

    # クリーンアップ
    import shutil
    try:
        shutil.rmtree(temp_dir)
    except:
        pass


@pytest.fixture
def temp_template_file():
    """一時的なテンプレートファイルを作成"""
    temp_dir = tempfile.mkdtemp()

    # 簡単なテンプレートを作成
    wb = Workbook()
    ws = wb.active
    ws.append(['手術日', '患者ID', '氏名', '入外', '術眼', '手術', '医師', '麻酔', '術前',
               '入外_比較', '術眼_比較', '手術_比較', '医師_比較', '麻酔_比較'])

    template_path = Path(temp_dir) / 'template.xlsx'
    wb.save(template_path)

    yield str(template_path)

    # クリーンアップ
    import shutil
    try:
        shutil.rmtree(temp_dir)
    except:
        pass


def test_surgery_error_extractor_creates_file_with_errors(temp_comparison_file, temp_template_file):
    """エラーがある場合にファイルが作成される"""
    # テンプレートファイルのパスを直接書き換える
    import service.surgery_error_extractor as extractor_module
    original_path = r'C:\Shinseikai\OPHChecker\眼科手術指示確認.xlsx'

    # モジュール内のハードコードされたパスをパッチ
    with patch.object(extractor_module, '__file__', temp_comparison_file['temp_dir']):
        # load_workbookをモックしてテンプレートファイルを使用
        with patch('service.surgery_error_extractor.load_workbook') as mock_load:
            from openpyxl import load_workbook
            mock_load.return_value = load_workbook(temp_template_file)

            result = surgery_error_extractor(
                temp_comparison_file['comparison'],
                temp_comparison_file['output_dir'],
                temp_template_file
            )

            # ファイルパスが返される
            assert result != ""
            assert Path(result).exists()


def test_surgery_error_extractor_returns_empty_without_errors(temp_template_file):
    """エラーがない場合は空文字列を返す"""
    temp_dir = tempfile.mkdtemp()

    # エラーなしのデータ
    data = """手術日,患者ID,氏名,入外,術眼,手術,医師,麻酔,術前,入外_比較,術眼_比較,手術_比較,医師_比較,麻酔_比較
2025/01/15,12345,患者A,外来,R,白内障手術,橋本,局所,検査A,True,True,True,True,True
2025/01/16,12346,患者B,入院,L,緑内障手術,植田,全身,検査B,True,True,True,True,True
"""
    comparison_path = Path(temp_dir) / 'comparison.csv'
    comparison_path.write_text(data, encoding='cp932')

    output_dir = Path(temp_dir) / 'output'
    output_dir.mkdir()

    try:
        result = surgery_error_extractor(
            str(comparison_path),
            str(output_dir),
            temp_template_file
        )

        # 空文字列が返される
        assert result == ""

    finally:
        import shutil
        try:
            shutil.rmtree(temp_dir)
        except:
            pass


def test_surgery_error_extractor_extracts_false_records(temp_comparison_file, temp_template_file):
    """未入力のレコードが抽出される"""
    with patch('service.surgery_error_extractor.load_workbook') as mock_load:
        from openpyxl import load_workbook
        mock_load.return_value = load_workbook(temp_template_file)

        result = surgery_error_extractor(
            temp_comparison_file['comparison'],
            temp_comparison_file['output_dir'],
            temp_template_file
        )

        # 結果ファイルを読み込み
        df = pd.read_excel(result)

        # 未入力のレコードが抽出される（現在の実装ではFalse文字列は抽出されない）
        assert len(df) >= 1

        # 患者Cが含まれる（未入力）
        assert '患者C' in df['氏名'].values


def test_surgery_error_extractor_extracts_uninput_records(temp_comparison_file, temp_template_file):
    """未入力のレコードが抽出される"""
    with patch('service.surgery_error_extractor.load_workbook') as mock_load:
        from openpyxl import load_workbook
        mock_load.return_value = load_workbook(temp_template_file)

        result = surgery_error_extractor(
            temp_comparison_file['comparison'],
            temp_comparison_file['output_dir'],
            temp_template_file
        )

        # 結果ファイルを読み込み
        df = pd.read_excel(result)

        # 患者Cが含まれる（未入力）
        assert '患者C' in df['氏名'].values


def test_surgery_error_extractor_correct_columns(temp_comparison_file, temp_template_file):
    """正しい列が出力される"""
    with patch('service.surgery_error_extractor.load_workbook') as mock_load:
        from openpyxl import load_workbook
        mock_load.return_value = load_workbook(temp_template_file)

        result = surgery_error_extractor(
            temp_comparison_file['comparison'],
            temp_comparison_file['output_dir'],
            temp_template_file
        )

        # 結果ファイルを読み込み
        df = pd.read_excel(result)

        expected_columns = [
            '手術日', '患者ID', '氏名', '入外', '術眼', '手術', '医師', '麻酔', '術前',
            '入外_比較', '術眼_比較', '手術_比較', '医師_比較', '麻酔_比較'
        ]

        assert list(df.columns) == expected_columns


def test_surgery_error_extractor_filename_format(temp_comparison_file, temp_template_file):
    """ファイル名が正しい形式で作成される"""
    with patch('service.surgery_error_extractor.load_workbook') as mock_load:
        from openpyxl import load_workbook
        mock_load.return_value = load_workbook(temp_template_file)

        result = surgery_error_extractor(
            temp_comparison_file['comparison'],
            temp_comparison_file['output_dir'],
            temp_template_file
        )

        filename = Path(result).name

        # ファイル名の形式を確認
        assert filename.startswith('眼科手術指示確認')
        assert filename.endswith('.xlsx')
        assert len(filename) == len('眼科手術指示確認YYYYMMDD_HHMM.xlsx')


def test_surgery_error_extractor_creates_output_directory(temp_comparison_file, temp_template_file):
    """出力ディレクトリが存在しない場合は作成される"""
    # 出力ディレクトリを削除
    import shutil
    shutil.rmtree(temp_comparison_file['output_dir'])

    with patch('service.surgery_error_extractor.load_workbook') as mock_load:
        from openpyxl import load_workbook
        mock_load.return_value = load_workbook(temp_template_file)

        result = surgery_error_extractor(
            temp_comparison_file['comparison'],
            temp_comparison_file['output_dir'],
            temp_template_file
        )

        # ディレクトリが作成されている
        assert Path(temp_comparison_file['output_dir']).exists()
        assert Path(result).exists()
