import logging
from datetime import datetime
from pathlib import Path
from typing import cast

import pandas as pd
from openpyxl import load_workbook


def surgery_error_extractor(comparison_result: str, output_path: str, template_path: str) -> str:
    """
    comparison_resultからFALSEまたは未入力が含まれる行を抽出し眼科手術指示確認.xlsxとして出力

    Args:
        comparison_result: 比較結果CSVファイルのパス
        output_path: 出力先ディレクトリのパス
        template_path: テンプレートExcelファイルのパス

    Returns:
        生成されたファイルのパス
    """
    df = pd.read_csv(comparison_result, encoding='cp932')
    comparison_cols = ['入外_比較', '術眼_比較', '手術_比較', '医師_比較', '麻酔_比較']

    # FALSEまたは「未入力」が含まれる行を抽出
    mask = pd.Series([False] * len(df))
    for col in comparison_cols:
        mask |= (df[col] == False) | (df[col] == '未入力')

    df_errors = df[mask]

    if len(df_errors) == 0:
        logging.info("不一致および未入力はありませんでした")
        return ""

    output_columns = ['手術日', '患者ID', '氏名', '入外', '術眼', '手術', '医師', '麻酔', '術前','入外_比較', '術眼_比較', '手術_比較', '医師_比較', '麻酔_比較']
    df_output = cast(pd.DataFrame, df_errors[output_columns].copy())

    Path(output_path).mkdir(parents=True, exist_ok=True)

    timestamp = datetime.now().strftime('%Y%m%d_%H%M')
    output_filename = f'眼科手術指示確認{timestamp}.xlsx'
    output_filepath = Path(output_path) / output_filename

    wb = load_workbook(template_path)
    ws = wb.active

    for row_idx, (_, row_data) in enumerate(df_output.iterrows(), start=2):
        for col_idx, value in enumerate(row_data, start=1):
            if ws is not None:
                # 手術日列（1列目）の場合、datetimeオブジェクトに変換
                if col_idx == 1 and pd.notna(value):
                    try:
                        # 文字列をdatetimeオブジェクトに変換
                        value = datetime.strptime(str(value), '%Y/%m/%d')
                    except (ValueError, TypeError):
                        # 変換できない場合はそのまま設定
                        pass
                
                ws.cell(row=row_idx, column=col_idx).value = value

    wb.save(output_filepath)

    logging.info(f"眼科手術指示確認ファイルの作成が完了しました")
    logging.info(f"エラー件数: {len(df_errors)}件")
    logging.info(f"出力ファイル: {output_filepath}")

    return str(output_filepath)


if __name__ == '__main__':
    from utils.config_manager import get_paths, load_config

    config = load_config()
    paths = get_paths(config)

    surgery_error_extractor(
        paths['comparison_result'],
        paths['output_path'],
        paths['template_path']
    )
