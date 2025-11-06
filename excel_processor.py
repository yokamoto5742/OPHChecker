from openpyxl import load_workbook
from openpyxl.styles import Alignment


def process_excel(input_file: str, output_file: str) -> None:
    """
    Excelファイルを処理して、指定された操作を実行します。
    
    Args:
        input_file: 入力Excelファイルのパス
        output_file: 出力Excelファイルのパス
    """
    wb = load_workbook(input_file)
    sheet = wb.active
    
    # 1行目を削除
    sheet.delete_rows(1)
    
    # B列からG列を削除（元のB:G = 2:7列目）
    sheet.delete_cols(2, 6)
    
    # D列からE列を削除（削除後の新しいD:E = 4:5列目）
    sheet.delete_cols(4, 2)
    
    # H列からN列を削除（削除後の新しいH:N = 8:14列目）
    sheet.delete_cols(8, 7)
    
    # A列からG列の列幅を12に設定
    for col in range(1, 8):
        column_letter = sheet.cell(row=1, column=col).column_letter
        sheet.column_dimensions[column_letter].width = 12
    
    # A列からG列の配置を中央揃えに設定
    center_alignment = Alignment(horizontal='center', vertical='center')
    for row in sheet.iter_rows(min_col=1, max_col=7):
        for cell in row:
            cell.alignment = center_alignment
    
    # F列の文字列を置換
    for cell in sheet['F']:
        if cell.value and isinstance(cell.value, str):
            cell.value = cell.value.replace('(ｸﾗﾚｵﾝﾄｰﾘｯｸ)', '')
    
    wb.save(output_file)
    print(f"処理完了: {output_file}")


if __name__ == '__main__':
    input_path = 'C:\Shinseikai\OPHChecker\手術予定表.xls'
    output_path = 'C:\Shinseikai\OPHChecker\processed_surgery_schedule.csv'
    
    process_excel(input_path, output_path)
