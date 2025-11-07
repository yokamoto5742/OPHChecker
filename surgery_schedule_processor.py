import os
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment


def process_excel(input_file: str, output_file: str, sheet_name: str = '南館2') -> None:
    """
    Excelファイルを処理して、指定された操作を実行します。
    .xls、.xlsx形式の入力に対応し、.xlsx、.csv形式での出力に対応します。

    Args:
        input_file: 入力Excelファイルのパス（.xls または .xlsx）
        output_file: 出力ファイルのパス（.xlsx または .csv）
        sheet_name: 読み込むシート名（デフォルト: '南館2'）
    """
    input_ext = Path(input_file).suffix.lower()
    output_ext = Path(output_file).suffix.lower()

    # .xlsファイルの場合、pandasで読み込んでから一時的に.xlsxとして保存
    if input_ext == '.xls':
        print(f".xlsファイルを検出しました。シート '{sheet_name}' を.xlsx形式に変換中...")
        df_temp = pd.read_excel(input_file, sheet_name=sheet_name, engine='xlrd')
        temp_xlsx = input_file.rsplit('.', 1)[0] + '_temp.xlsx'
        df_temp.to_excel(temp_xlsx, index=False)
        working_file = temp_xlsx
        is_temp_file = True
    else:
        working_file = input_file
        is_temp_file = False

    try:
        # openpyxlでファイルを読み込み
        wb = load_workbook(working_file)

        # .xlsxファイルの場合、指定されたシート名を取得
        if input_ext == '.xlsx':
            if sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                print(f"シート '{sheet_name}' を読み込みました")
            else:
                raise ValueError(f"シート '{sheet_name}' が見つかりません。利用可能なシート: {wb.sheetnames}")
        else:
            # .xlsから変換した場合は既に指定されたシートのみが含まれている
            sheet = wb.active

        # 1行目を削除
        sheet.delete_rows(1)

        # B列からG列を削除（元のB:G = 2:7列目）
        sheet.delete_cols(2, 6)

        # D列からE列を削除（削除後の新しいD:E = 4:5列目）
        sheet.delete_cols(4, 2)

        # H列からN列を削除（削除後の新しいH:N = 8:14列目）
        sheet.delete_cols(8, 7)

        # A列からG列の列幅を12に設定
        for col in range(1, 8):
            column_letter = sheet.cell(row=1, column=col).column_letter
            sheet.column_dimensions[column_letter].width = 12

        # A列からG列の配置を中央揃えに設定
        center_alignment = Alignment(horizontal='center', vertical='center')
        for row in sheet.iter_rows(min_col=1, max_col=7):
            for cell in row:
                cell.alignment = center_alignment

        # F列の文字列を置換
        for cell in sheet['F']:
            if cell.value and isinstance(cell.value, str):
                cell.value = cell.value.replace('(ｸﾗﾚｵﾝﾄｰﾘｯｸ)', '')

        # 出力形式に応じて保存
        if output_ext == '.csv':
            # CSVとして保存する場合、pandasを使用
            print(f"CSV形式で保存中...")
            data = []
            for row in sheet.iter_rows(values_only=True):
                data.append(row)
            df_output = pd.DataFrame(data)
            df_output.to_csv(output_file, index=False, header=False, encoding='utf-8-sig')
        else:
            # .xlsxとして保存
            wb.save(output_file)

        print(f"処理完了: {output_file}")

    finally:
        # 一時ファイルを削除
        if is_temp_file and os.path.exists(working_file):
            os.remove(working_file)
            print(f"一時ファイルを削除しました: {working_file}")


if __name__ == '__main__':
    input_path = r'C:\Shinseikai\OPHChecker\input\手術予定表.xls'
    output_path = r'C:\Shinseikai\OPHChecker\processed_surgery_schedule.csv'

    process_excel(input_path, output_path)
